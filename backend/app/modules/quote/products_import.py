"""
products_import.py — Importación masiva de productos vía Excel (módulo QUOTE).

Permite poblar el catálogo de productos de una sola vez desde un archivo .xlsx con
columnas fijas (ID, CATEGORIA, NOMBRE, DESCRIPCION, MARCA, PRECIO, UNIDAD, MONEDA),
pensado para negocios que ya tienen su catálogo en Excel y prefieren cargarlo en
bloque en vez de crear producto por producto.

La columna ID es solo referencial: el id real lo asigna la base de datos (SERIAL),
así el archivo puede traer cualquier numeración interna del cliente sin chocar con
la secuencia de la tabla `productos`.
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError

from app.modules.quote.schemas import ProductoCreate

# Encabezados de la plantilla, en orden. "ID" es solo referencial (se ignora al importar).
HEADERS = ["ID", "CATEGORIA", "NOMBRE", "DESCRIPCION", "MARCA", "PRECIO", "UNIDAD", "MONEDA"]
REQUIRED_HEADERS = ["CATEGORIA", "NOMBRE", "DESCRIPCION", "MARCA", "PRECIO", "UNIDAD", "MONEDA"]
MONEDAS_VALIDAS = {"PEN", "USD"}
MAX_ROWS = 5000


class ImportStructureError(Exception):
    """Error de estructura del archivo (encabezados faltantes, archivo vacío/corrupto,
    demasiadas filas). A diferencia de los errores de validación por fila, estos
    impiden leer el archivo por completo."""


def generate_template_xlsx() -> bytes:
    """Genera la plantilla .xlsx en blanco: encabezados, una fila de ejemplo y un
    desplegable de MONEDA (PEN/USD) para evitar errores de tipeo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ejemplo = ["1", "Panel Solar", "Panel Solar Jinko 555Wp Monofacial",
               "Panel solar monofacial 555Wp", "Jinko Solar", "154.88", "Und", "USD"]
    for col, valor in enumerate(ejemplo, start=1):
        ws.cell(row=2, column=col, value=valor)

    moneda_col = HEADERS.index("MONEDA") + 1
    dv = DataValidation(type="list", formula1='"PEN,USD"', allow_blank=True, showErrorMessage=True)
    dv.error = "La moneda debe ser PEN o USD"
    ws.add_data_validation(dv)
    col_letter = ws.cell(row=2, column=moneda_col).column_letter
    dv.add(f"{col_letter}2:{col_letter}{MAX_ROWS + 1}")

    widths = [6, 22, 40, 40, 18, 12, 10, 10]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_str(value) -> str:
    return str(value).strip() if value is not None else ""


def parse_products_xlsx(raw: bytes) -> tuple[list[ProductoCreate], list[dict]]:
    """Lee la primera hoja del archivo y devuelve (productos_validos, errores_por_fila).

    - Errores de ESTRUCTURA (archivo corrupto, sin encabezados, demasiadas filas)
      levantan ImportStructureError: no se pudo ni empezar a leer el catálogo.
    - Errores de VALIDACIÓN por fila (precio inválido, campo vacío, moneda no
      reconocida) se acumulan en `errores` en vez de interrumpir la lectura, para
      poder reportar TODAS las filas problemáticas de una sola vez.
    """
    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise ImportStructureError(f"No se pudo leer el archivo Excel: {e}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ImportStructureError("El archivo está vacío.")

    header_map: dict[str, int] = {}
    for idx, raw_header in enumerate(header_row or []):
        name = _cell_str(raw_header).upper()
        if name:
            header_map[name] = idx

    faltantes = [h for h in REQUIRED_HEADERS if h not in header_map]
    if faltantes:
        raise ImportStructureError(
            "Faltan columnas obligatorias en la plantilla: " + ", ".join(faltantes)
        )

    productos: list[ProductoCreate] = []
    errores: list[dict] = []

    fila_num = 1  # la fila 1 es el encabezado
    filas_datos = 0
    for row in rows_iter:
        fila_num += 1
        if row is None or all(_cell_str(v) == "" for v in row):
            continue  # fila en blanco: se ignora en silencio

        filas_datos += 1
        if filas_datos > MAX_ROWS:
            raise ImportStructureError(f"El archivo supera el máximo de {MAX_ROWS} filas por importación.")

        def cell(col: str):
            idx = header_map.get(col)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        precio_cell = cell("PRECIO")
        if isinstance(precio_cell, (int, float)):
            precio = Decimal(str(precio_cell))
        else:
            precio_str = _cell_str(precio_cell)
            try:
                precio = Decimal(precio_str) if precio_str else Decimal("0")
            except InvalidOperation:
                errores.append({"fila": fila_num, "errores": [f"Precio inválido: '{precio_str}'"]})
                continue

        moneda = _cell_str(cell("MONEDA")).upper()
        fila_errores: list[str] = []
        if moneda and moneda not in MONEDAS_VALIDAS:
            fila_errores.append(f"Moneda inválida: '{moneda}' (debe ser PEN o USD)")
        if fila_errores:
            errores.append({"fila": fila_num, "errores": fila_errores})
            continue

        payload = {
            "categoria": _cell_str(cell("CATEGORIA")),
            "nombre": _cell_str(cell("NOMBRE")),
            "descripcion": _cell_str(cell("DESCRIPCION")),
            "marca": _cell_str(cell("MARCA")),
            "unidad": _cell_str(cell("UNIDAD")) or "Und",
            "precio": precio,
            "moneda": moneda or "PEN",
        }

        try:
            productos.append(ProductoCreate(**payload))
        except ValidationError as e:
            mensajes = [err["msg"].replace("Value error, ", "") for err in e.errors()]
            errores.append({"fila": fila_num, "errores": mensajes})

    if not productos and not errores:
        raise ImportStructureError("El archivo no contiene filas de datos para importar.")

    return productos, errores
